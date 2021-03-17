# -*- coding : UTF-8 -*-
import xlrd
import os
import time
from openpyxl import load_workbook
import easygui as g
import sys
import os
from tqdm import tqdm

class OpenExl():
    #初始化exl文件，by_index表示第几个工作表单，默认值为0。
    def __init__(self, exc_file, filename, by_index = 0):
        self.datalist = []
        self.count = 1
        self.by_index = by_index
        self.exc_file = exc_file
        self.filename = filename
        # 打开文件
        try:
            self.exc_data = xlrd.open_workbook(self.exc_file)
        except FileNotFoundError as fnf:
            print('EXCEL文件未找到，请再次确认输入的文件名')
            # sys.exit(1)
        self.wb = load_workbook(self.filename)
       # print(self.wb.sheetnames)
        self.ws = self.wb['Sheet1']
        self.ws.delete_rows(2,1999)
        self.table = self.exc_data.sheets()[self.by_index]
        # time.sleep(5)


    # 定义获取excle表单行数和列表
    def getCol_Row(self):
        self.nrows_row = self.table.nrows  # 行数
        self.nrows_col = self.table.ncols  # 列数
        return self.nrows_row, self.nrows_col

    def getYear_Month(self, rownum = 0):
        '''
        处理数据，获得年月时间
        :return: 返回年月
        '''
        self.year_month = self.table.row_values(rownum)[0].split(' ')[3]
        self.year_month = self.year_month[:self.year_month.rfind('-')]
        self.year_month = self.year_month.replace('-', '/')

        return self.year_month

    def getDay(self, rownum = 2):
        '''
        获得日期列表
        :param rownum: 默认从第三行获取日期数据
        :return:日期列表
        '''
        self.day = []
        day_num = self.table.row_values(rownum)
        # print(day_num)
        for oneday in day_num:
            if oneday.isdecimal():
                self.day.append(oneday)
            else:
                # oneday = ''
                self.day.append('')

        return self.day

    def getAttence(self, rownum = 3):
        '''
        返回考勤统计数据列表
        :param rownum:默认从第三行开始
        :return:self.datalist
        '''
        #获得改行excle的所有数据
        self.attence_data = self.table.row_values(rownum)
        #获得用户姓名
        self.name = self.attence_data[0]
        #获得用户编号
        self.userid = self.attence_data[5]
        for onceday in self.day:
            #判断日期不为空，并且日子字符里是10进制数字
            if onceday != '' and onceday.isdecimal():
                am = self.year_month + '/' + onceday + '  ' + '08:30:00'
                noon = self.year_month + '/' + onceday + '  ' + '14:00:00'
                pm = self.year_month + '/' + onceday + '  ' + '17:30:00'
                befor_pm = self.year_month + '/' + onceday + '  ' + '16:30:00'
                # print(am, noon, pm)
                #获得日期的列表索引
                num = self.day.index(onceday)
                # print(self.attence_data[num] + '\n')
                #过滤获得的考勤数据
                attence_time = self.attence_data[num].replace('外勤', ' ').replace('  ', ' ').replace('\n', '').rstrip().split(' ')
                # print(attence_time)
                if attence_time != ['']:
                    # print(onceday, attence_time)
                    #循环列表数据，判断考勤数据情况
                    for once_time in attence_time:
                        data = []
                        once_time = self.year_month + '/' + onceday + '  ' + once_time + ':00'

                        data.append(self.count)
                        data.append(self.name)
                        data.append(self.userid)
                        data.append(self.userid)
                        data.append('刷卡考勤')
                        data.append('成功')

                        if comTime(once_time, am) != 3 and comTime(once_time, noon) == 1:
                            data.append('进')
                            data.append('')
                        elif comTime(once_time, am) == 3 and comTime(once_time, noon) == 1:
                            #早上迟到
                            data.append('进')
                            data.append('早上迟到')
                        elif comTime(once_time, noon) == 1 and comTime(once_time, befor_pm) == 1:
                            data.append('进')
                            data.append('')
                        elif comTime(once_time, noon) == 3 and comTime(once_time, befor_pm) == 1:
                            #下午迟到
                            data.append('进')
                            data.append('下午迟到')
                        elif comTime(once_time, befor_pm) == 3 and comTime(once_time, pm) == 1:
                            #下午早退
                            data.append('出')
                            data.append('下午早退')
                        elif comTime(once_time, pm) != 1:
                            data.append('出')
                            data.append('')
                        data.append(once_time)
                        data.append('1')
                        self.datalist.append(data)
                        self.count += 1

                else:
                    data = []
                    once_time = self.year_month + '/' + onceday

                    data.append(self.count)
                    data.append(self.name)
                    data.append(self.userid)
                    data.append(self.userid)
                    data.append('')
                    data.append('未打卡')
                    data.append('未打卡')
                    data.append('')
                    data.append(once_time)
                    data.append('1')
                    self.datalist.append(data)
                    self.count += 1

        # print(self.datalist)
        return self.datalist


    def write_data(self, data):
        '''
        定义保存Excel函数
        Args:
            data:保存的数据列表
        '''
        self.ws.append(data)
        self.wb.save(self.filename)
        # self.wb.close()

def comTime(begintime = '', endtime = ''):
        time1 = begintime
        # print(time1)
        time2 = endtime
        # print(time2)
        t1 = time.mktime(time.strptime(time1, "%Y/%m/%d  %H:%M:%S"))
        t2 = time.mktime(time.strptime(time2, "%Y/%m/%d  %H:%M:%S"))
        if t1 < t2:
            time_flag = 1
        elif t1 == t2:
            time_flag = 2
        else:
            time_flag = 3
        # print(time_flag)
        return time_flag


def getExc_File(msg):
    try:
        exc_filename = g.enterbox(msg, '输入EXCEL表格文件名')
        print('输入的排查EXCEL文件名为：',exc_filename)
    except Exception:
        sys.exit(1)
    else:
        if exc_filename == '':
            msg = '【请输入正确EXCEL表格文件名】，文件格式请使用xlsx'
            getExc_File(msg)
    return exc_filename

def getWay():
    myway = os.getcwd()
    # print(myway)
    return myway

if __name__ == "__main__":
    msg = '请输入需要排查的EXCEL表格文件名，文件格式请使用xlsx'
    thePath = getWay()
    exc_filename = getExc_File(msg)
    exc_file = thePath + "\\" + exc_filename
    sava_file = thePath + "\打卡统计.xlsx"
    # print(exc_file ,sava_file)
    oe = OpenExl(exc_file, sava_file)
    year_month = oe.getYear_Month()
    print('统计考勤时间为：',year_month)
    day = oe.getDay()
    # print(day)
    row_num = oe.getCol_Row()
    # print(row_num)
    print('\n---------->>>考勤统计开始，期间请不要打开考勤Excel文件<<<----------')
    for row in tqdm(range(3, row_num[0])):
        attence = oe.getAttence(row)
        for data in attence:
            oe.write_data(data)
        oe.datalist = []
        row += 1

    oe.wb.close()


